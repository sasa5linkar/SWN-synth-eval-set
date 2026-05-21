from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from itertools import combinations
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook


TACNO = "ta\u010dno"
BLIZU = "blizu"
NETACNO = "neta\u010dno"
STATUS_ORDER = [NETACNO, BLIZU, TACNO]
STATUS_DISPLAY_ORDER = [TACNO, BLIZU, NETACNO]

POSITIVE_VALUES = [
    "nije pozitivan",
    "slabo pozitivan",
    "pozitivan",
    "veoma pozitivan",
    "ekstremno pozitivan",
]
NEGATIVE_VALUES = [
    "nije negativan",
    "slabo negativan",
    "negativan",
    "veoma negativan",
    "ekstremno negativan",
]
VALUE_SCORES = {
    "positive": {value: index * 0.25 for index, value in enumerate(POSITIVE_VALUES)},
    "negative": {value: index * 0.25 for index, value in enumerate(NEGATIVE_VALUES)},
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.lower() if text else ""


def pct(numerator: int | float, denominator: int | float) -> float:
    return round((numerator / denominator * 100), 1) if denominator else 0.0


def counter_dict(counter: Counter) -> dict[str, int]:
    return {str(key): value for key, value in counter.items()}


def ordered_counter_dict(counter: Counter, order: list[str]) -> dict[str, int]:
    return {key: counter.get(key, 0) for key in order if counter.get(key, 0)}


def load_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Evaluation" not in wb.sheetnames:
        raise ValueError(f"{path} has no Evaluation sheet")
    ws = wb["Evaluation"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    ]
    return headers, rows


def effective_value(row: dict[str, Any], axis: str) -> str:
    status = norm(row.get(f"human_{axis}_eval"))
    ollama_value = norm(row.get(f"ollama_{axis}_value"))
    human_value = norm(row.get(f"human_{axis}_value"))
    return ollama_value if status == TACNO or not human_value else human_value


def axis_stats(rows: list[dict[str, Any]], axis: str) -> dict[str, Any]:
    eval_col = f"human_{axis}_eval"
    human_col = f"human_{axis}_value"
    ollama_col = f"ollama_{axis}_value"
    eval_counts = Counter(norm(row.get(eval_col)) for row in rows)
    ollama_counts = Counter(norm(row.get(ollama_col)) for row in rows)
    effective_counts = Counter()
    corrections: list[dict[str, str]] = []
    distance_counts: Counter = Counter()
    missing_values: list[str] = []
    unchanged_corrections: list[str] = []

    for row in rows:
        status = norm(row.get(eval_col))
        ollama_value = norm(row.get(ollama_col))
        human_value = norm(row.get(human_col))
        final_value = effective_value(row, axis)
        effective_counts[final_value] += 1

        if status in (BLIZU, NETACNO):
            if not human_value:
                missing_values.append(row.get("ILI"))
            else:
                corrections.append(
                    {
                        "ILI": row.get("ILI"),
                        "status": status,
                        "ollama_value": ollama_value,
                        "human_value": human_value,
                    }
                )
                if ollama_value == human_value:
                    unchanged_corrections.append(row.get("ILI"))
                scores = VALUE_SCORES[axis]
                if ollama_value in scores and human_value in scores:
                    distance = abs(scores[ollama_value] - scores[human_value])
                    distance_counts[str(distance)] += 1

    return {
        "eval_counts": ordered_counter_dict(eval_counts, STATUS_DISPLAY_ORDER),
        "exact_count": eval_counts[TACNO],
        "near_count": eval_counts[BLIZU],
        "wrong_count": eval_counts[NETACNO],
        "exact_rate_pct": pct(eval_counts[TACNO], len(rows)),
        "exact_or_near_rate_pct": pct(eval_counts[TACNO] + eval_counts[BLIZU], len(rows)),
        "ollama_value_counts": counter_dict(ollama_counts),
        "effective_human_value_counts": counter_dict(effective_counts),
        "corrections_count": len(corrections),
        "correction_distance_counts": counter_dict(distance_counts),
        "missing_human_values": missing_values,
        "unchanged_corrections": unchanged_corrections,
        "corrections": corrections,
    }


def single_stats(label: str, source: Path, out_root: Path) -> dict[str, Any]:
    out_dir = out_root / label
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / f"human_eval_{label}.xlsx"
    out_csv = out_dir / f"human_eval_{label}.csv"
    out_json = out_dir / f"human_eval_{label}_stats.json"
    out_md = out_dir / "README.md"

    copy2(source, out_xlsx)
    headers, rows = load_rows(out_xlsx)

    with out_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header) for header in headers})

    positive = axis_stats(rows, "positive")
    negative = axis_stats(rows, "negative")
    row_quality = Counter()
    status_pairs = Counter()
    all_statuses: list[str] = []
    quality_flags: list[dict[str, Any]] = []

    for row in rows:
        pos = norm(row.get("human_positive_eval"))
        neg = norm(row.get("human_negative_eval"))
        status_pairs[(pos, neg)] += 1
        all_statuses.extend([pos, neg])
        if pos == TACNO and neg == TACNO:
            row_quality["both_exact"] += 1
        elif pos in (TACNO, BLIZU) and neg in (TACNO, BLIZU):
            row_quality["acceptable_with_near"] += 1
        elif pos == NETACNO or neg == NETACNO:
            row_quality["has_wrong_axis"] += 1
        else:
            row_quality["other"] += 1

    for axis, stats in (("positive", positive), ("negative", negative)):
        if stats["missing_human_values"]:
            quality_flags.append(
                {
                    "type": "missing_human_value_for_correction",
                    "axis": axis,
                    "ili": stats["missing_human_values"],
                }
            )
        if stats["unchanged_corrections"]:
            quality_flags.append(
                {
                    "type": "correction_status_but_same_value",
                    "axis": axis,
                    "ili": stats["unchanged_corrections"],
                }
            )

    axis_total_counts = Counter(all_statuses)
    stats = {
        "label": f"human_eval_{label}",
        "source_file": str(source),
        "copied_workbook": str(out_xlsx),
        "evaluation_csv": str(out_csv),
        "sheet": "Evaluation",
        "rows": len(rows),
        "judgements_total": len(rows) * 2,
        "model_counts": counter_dict(Counter(norm(row.get("ollama_model")) for row in rows)),
        "sentiment_swn_counts": counter_dict(Counter(norm(row.get("sentiment_SWN")) for row in rows)),
        "sentiment_lexicon_counts": counter_dict(Counter(norm(row.get("sentiment_lexicon")) for row in rows)),
        "axis_total_eval_counts": ordered_counter_dict(axis_total_counts, STATUS_DISPLAY_ORDER),
        "axis_total_exact_rate_pct": pct(axis_total_counts[TACNO], len(rows) * 2),
        "axis_total_exact_or_near_rate_pct": pct(
            axis_total_counts[TACNO] + axis_total_counts[BLIZU], len(rows) * 2
        ),
        "row_quality_counts": counter_dict(row_quality),
        "row_exact_rate_pct": pct(row_quality["both_exact"], len(rows)),
        "row_exact_or_near_rate_pct": pct(
            row_quality["both_exact"] + row_quality["acceptable_with_near"], len(rows)
        ),
        "positive": positive,
        "negative": negative,
        "positive_negative_eval_pair_counts": {
            f"{pos}|{neg}": count for (pos, neg), count in status_pairs.items()
        },
        "quality_flags": quality_flags,
    }

    out_json.write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_md.write_text(single_readme(label, source, stats), encoding="utf-8")
    return stats


def single_readme(label: str, source: Path, stats: dict[str, Any]) -> str:
    row_counts = stats["row_quality_counts"]
    axis_counts = Counter(stats["axis_total_eval_counts"])
    rows = stats["rows"]
    judgements = stats["judgements_total"]
    positive = stats["positive"]
    negative = stats["negative"]

    text = f"""# Human evaluation {label}

Source workbook: `{source.name}`  
Copied workbook: `human_eval_{label}.xlsx`  
Evaluation sheet rows: {rows}  
Model: `mistral-small3.2:latest`

## Summary

Each row has two human checks: one for the positive Ollama score and one for the negative Ollama score.

| Metric | Count | Rate |
|---|---:|---:|
| Rows where both axes are `{TACNO}` | {row_counts.get("both_exact", 0)} / {rows} | {stats["row_exact_rate_pct"]}% |
| Rows where both axes are `{TACNO}` or `{BLIZU}` | {row_counts.get("both_exact", 0) + row_counts.get("acceptable_with_near", 0)} / {rows} | {stats["row_exact_or_near_rate_pct"]}% |
| Rows with at least one `{NETACNO}` axis | {row_counts.get("has_wrong_axis", 0)} / {rows} | {pct(row_counts.get("has_wrong_axis", 0), rows)}% |
| Individual axis judgements that are `{TACNO}` | {axis_counts[TACNO]} / {judgements} | {stats["axis_total_exact_rate_pct"]}% |
| Individual axis judgements that are `{TACNO}` or `{BLIZU}` | {axis_counts[TACNO] + axis_counts[BLIZU]} / {judgements} | {stats["axis_total_exact_or_near_rate_pct"]}% |

## Axis detail

| Axis | `{TACNO}` | `{BLIZU}` | `{NETACNO}` | `{TACNO}` rate | `{TACNO}` + `{BLIZU}` rate |
|---|---:|---:|---:|---:|---:|
| Positive | {positive["exact_count"]} | {positive["near_count"]} | {positive["wrong_count"]} | {positive["exact_rate_pct"]}% | {positive["exact_or_near_rate_pct"]}% |
| Negative | {negative["exact_count"]} | {negative["near_count"]} | {negative["wrong_count"]} | {negative["exact_rate_pct"]}% | {negative["exact_or_near_rate_pct"]}% |

## Effective human value distribution

| Positive value | Count |
|---|---:|
"""
    for value, count in positive["effective_human_value_counts"].items():
        text += f"| `{value}` | {count} |\n"
    text += "\n| Negative value | Count |\n|---|---:|\n"
    for value, count in negative["effective_human_value_counts"].items():
        text += f"| `{value}` | {count} |\n"

    flags = stats["quality_flags"]
    if flags:
        text += "\n## Quality flags\n\n"
        for flag in flags:
            text += f"- `{flag['type']}` on `{flag['axis']}`: {', '.join(flag['ili'])}\n"

    text += f"""

Full machine-readable stats are in `human_eval_{label}_stats.json`; the extracted `Evaluation` sheet is in `human_eval_{label}.csv`.
"""
    return text


def cohen_kappa(left: list[str], right: list[str], labels: list[str]) -> float | None:
    n = len(left)
    if n == 0:
        return None
    observed = sum(1 for a, b in zip(left, right) if a == b) / n
    left_counts = Counter(left)
    right_counts = Counter(right)
    expected = sum(left_counts[label] * right_counts[label] for label in labels) / (n * n)
    if expected == 1:
        return None
    return round((observed - expected) / (1 - expected), 4)


def weighted_kappa(left: list[str], right: list[str], labels: list[str], quadratic: bool = True) -> float | None:
    n = len(left)
    if n == 0 or len(labels) < 2:
        return None
    index = {label: i for i, label in enumerate(labels)}
    max_distance = len(labels) - 1
    left_counts = Counter(left)
    right_counts = Counter(right)

    def weight(a: str, b: str) -> float:
        distance = abs(index[a] - index[b]) / max_distance
        return distance * distance if quadratic else distance

    observed = 0.0
    expected = 0.0
    for a, b in zip(left, right):
        if a in index and b in index:
            observed += weight(a, b)
    observed /= n
    for a in labels:
        for b in labels:
            expected += (left_counts[a] * right_counts[b] / (n * n)) * weight(a, b)
    if expected == 0:
        return None
    return round(1 - observed / expected, 4)


def confusion(left: list[str], right: list[str], labels: list[str]) -> dict[str, dict[str, int]]:
    return {
        a: {b: sum(1 for x, y in zip(left, right) if x == a and y == b) for b in labels}
        for a in labels
    }


def agreement_stats(left: list[str], right: list[str], labels: list[str]) -> dict[str, Any]:
    n = len(left)
    exact = sum(1 for a, b in zip(left, right) if a == b)
    return {
        "n": n,
        "exact_agreement_count": exact,
        "exact_agreement_pct": pct(exact, n),
        "cohen_kappa": cohen_kappa(left, right, labels),
        "quadratic_weighted_kappa": weighted_kappa(left, right, labels, quadratic=True),
        "confusion_matrix": confusion(left, right, labels),
    }


def compare_stats(label_a: str, label_b: str, out_root: Path) -> dict[str, Any]:
    workbook_a = out_root / label_a / f"human_eval_{label_a}.xlsx"
    workbook_b = out_root / label_b / f"human_eval_{label_b}.xlsx"
    _, rows_a = load_rows(workbook_a)
    _, rows_b = load_rows(workbook_b)
    by_ili_a = {row["ILI"]: row for row in rows_a}
    by_ili_b = {row["ILI"]: row for row in rows_b}
    common_ili = [row["ILI"] for row in rows_a if row["ILI"] in by_ili_b]
    only_a = sorted(set(by_ili_a) - set(by_ili_b))
    only_b = sorted(set(by_ili_b) - set(by_ili_a))

    out_dir = out_root / f"comparison_{label_a}_{label_b}"
    out_dir.mkdir(parents=True, exist_ok=True)
    all_csv = out_dir / f"comparison_{label_a}_{label_b}.csv"
    disagreement_csv = out_dir / f"disagreements_{label_a}_{label_b}.csv"
    out_json = out_dir / f"comparison_{label_a}_{label_b}_stats.json"
    out_md = out_dir / "README.md"

    detail_rows: list[dict[str, Any]] = []
    status_lists: dict[str, tuple[list[str], list[str]]] = {
        "all_axes": ([], []),
        "positive": ([], []),
        "negative": ([], []),
    }
    acceptable_lists: dict[str, tuple[list[str], list[str]]] = {
        "all_axes": ([], []),
        "positive": ([], []),
        "negative": ([], []),
    }
    value_lists: dict[str, tuple[list[str], list[str]]] = {
        "positive": ([], []),
        "negative": ([], []),
    }
    value_distances: dict[str, list[float]] = {"positive": [], "negative": []}
    row_status_both_axes_agree = 0
    row_values_both_axes_agree = 0
    row_acceptable_both_axes_agree = 0

    for ili in common_ili:
        row_a = by_ili_a[ili]
        row_b = by_ili_b[ili]
        detail: dict[str, Any] = {
            "ILI": ili,
            "lemma_names": row_a.get("lemma_names"),
            "definition": row_a.get("definition"),
            "sentiment_lexicon": row_a.get("sentiment_lexicon"),
            f"notes_{label_a}": row_a.get("human_notes"),
            f"notes_{label_b}": row_b.get("human_notes"),
        }
        axis_status_agree: list[bool] = []
        axis_value_agree: list[bool] = []
        axis_acceptable_agree: list[bool] = []

        for axis in ["positive", "negative"]:
            status_a = norm(row_a.get(f"human_{axis}_eval"))
            status_b = norm(row_b.get(f"human_{axis}_eval"))
            value_a = effective_value(row_a, axis)
            value_b = effective_value(row_b, axis)
            scores = VALUE_SCORES[axis]
            distance = (
                abs(scores[value_a] - scores[value_b])
                if value_a in scores and value_b in scores
                else None
            )
            acceptable_a = "acceptable" if status_a in (TACNO, BLIZU) else "not_acceptable"
            acceptable_b = "acceptable" if status_b in (TACNO, BLIZU) else "not_acceptable"

            status_lists[axis][0].append(status_a)
            status_lists[axis][1].append(status_b)
            status_lists["all_axes"][0].append(status_a)
            status_lists["all_axes"][1].append(status_b)
            acceptable_lists[axis][0].append(acceptable_a)
            acceptable_lists[axis][1].append(acceptable_b)
            acceptable_lists["all_axes"][0].append(acceptable_a)
            acceptable_lists["all_axes"][1].append(acceptable_b)
            value_lists[axis][0].append(value_a)
            value_lists[axis][1].append(value_b)
            if distance is not None:
                value_distances[axis].append(distance)

            axis_status_agree.append(status_a == status_b)
            axis_value_agree.append(value_a == value_b)
            axis_acceptable_agree.append(acceptable_a == acceptable_b)

            detail[f"{axis}_status_{label_a}"] = status_a
            detail[f"{axis}_status_{label_b}"] = status_b
            detail[f"{axis}_status_agree"] = status_a == status_b
            detail[f"{axis}_acceptable_{label_a}"] = acceptable_a
            detail[f"{axis}_acceptable_{label_b}"] = acceptable_b
            detail[f"{axis}_acceptable_agree"] = acceptable_a == acceptable_b
            detail[f"{axis}_value_{label_a}"] = value_a
            detail[f"{axis}_value_{label_b}"] = value_b
            detail[f"{axis}_value_agree"] = value_a == value_b
            detail[f"{axis}_value_distance"] = distance

        detail["row_status_both_axes_agree"] = all(axis_status_agree)
        detail["row_values_both_axes_agree"] = all(axis_value_agree)
        detail["row_acceptable_both_axes_agree"] = all(axis_acceptable_agree)
        row_status_both_axes_agree += int(detail["row_status_both_axes_agree"])
        row_values_both_axes_agree += int(detail["row_values_both_axes_agree"])
        row_acceptable_both_axes_agree += int(detail["row_acceptable_both_axes_agree"])
        detail_rows.append(detail)

    fieldnames = list(detail_rows[0].keys()) if detail_rows else []
    for csv_path, rows in (
        (all_csv, detail_rows),
        (
            disagreement_csv,
            [
                row
                for row in detail_rows
                if not row["row_status_both_axes_agree"]
                or not row["row_values_both_axes_agree"]
                or not row["row_acceptable_both_axes_agree"]
            ],
        ),
    ):
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    status_agreement = {
        key: agreement_stats(left, right, STATUS_ORDER)
        for key, (left, right) in status_lists.items()
    }
    acceptable_agreement = {
        key: agreement_stats(left, right, ["not_acceptable", "acceptable"])
        for key, (left, right) in acceptable_lists.items()
    }
    value_agreement: dict[str, Any] = {}
    for axis, labels in (("positive", POSITIVE_VALUES), ("negative", NEGATIVE_VALUES)):
        left, right = value_lists[axis]
        exact = sum(1 for a, b in zip(left, right) if a == b)
        within_one_step = sum(1 for distance in value_distances[axis] if distance <= 0.25)
        distances = value_distances[axis]
        value_agreement[axis] = {
            "n": len(left),
            "exact_agreement_count": exact,
            "exact_agreement_pct": pct(exact, len(left)),
            "within_one_step_count": within_one_step,
            "within_one_step_pct": pct(within_one_step, len(left)),
            "mean_absolute_distance": round(sum(distances) / len(distances), 4) if distances else None,
            "max_distance": max(distances) if distances else None,
            "quadratic_weighted_kappa": weighted_kappa(left, right, labels, quadratic=True),
            "confusion_matrix": confusion(left, right, labels),
        }

    disagreement_lists = {
        "positive_status": [
            row["ILI"] for row in detail_rows if not row["positive_status_agree"]
        ],
        "negative_status": [
            row["ILI"] for row in detail_rows if not row["negative_status_agree"]
        ],
        "positive_acceptable": [
            row["ILI"] for row in detail_rows if not row["positive_acceptable_agree"]
        ],
        "negative_acceptable": [
            row["ILI"] for row in detail_rows if not row["negative_acceptable_agree"]
        ],
        "positive_effective_value": [
            row["ILI"] for row in detail_rows if not row["positive_value_agree"]
        ],
        "negative_effective_value": [
            row["ILI"] for row in detail_rows if not row["negative_value_agree"]
        ],
    }
    any_disagreement = sorted(
        set().union(*(set(values) for values in disagreement_lists.values()))
    )
    full_agreement = [
        row["ILI"]
        for row in detail_rows
        if row["row_status_both_axes_agree"]
        and row["row_values_both_axes_agree"]
        and row["row_acceptable_both_axes_agree"]
    ]

    stats = {
        "comparison": f"{label_a}_vs_{label_b}",
        "common_rows": len(common_ili),
        "only_in_first": only_a,
        "only_in_second": only_b,
        "row_level": {
            "status_both_axes_agree_count": row_status_both_axes_agree,
            "status_both_axes_agree_pct": pct(row_status_both_axes_agree, len(common_ili)),
            "effective_values_both_axes_agree_count": row_values_both_axes_agree,
            "effective_values_both_axes_agree_pct": pct(row_values_both_axes_agree, len(common_ili)),
            "acceptable_both_axes_agree_count": row_acceptable_both_axes_agree,
            "acceptable_both_axes_agree_pct": pct(row_acceptable_both_axes_agree, len(common_ili)),
        },
        "status_agreement": status_agreement,
        "acceptable_binary_agreement": acceptable_agreement,
        "effective_value_agreement": value_agreement,
        "disagreement_summary": {
            "full_agreement_rows": full_agreement,
            "any_disagreement_rows": any_disagreement,
            "by_type": disagreement_lists,
        },
        "comparison_csv": str(all_csv),
        "disagreement_csv": str(disagreement_csv),
    }
    out_json.write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_md.write_text(comparison_readme(label_a, label_b, stats), encoding="utf-8")
    return stats


def kappa_text(value: float | None) -> str:
    return "" if value is None else str(value)


def comparison_readme(label_a: str, label_b: str, stats: dict[str, Any]) -> str:
    row = stats["row_level"]
    status = stats["status_agreement"]
    acceptable = stats["acceptable_binary_agreement"]
    values = stats["effective_value_agreement"]
    disagreements = stats["disagreement_summary"]
    text = f"""# Human evaluation comparison {label_a} vs {label_b}

Compared rows: {stats["common_rows"]}

## Recommended agreement statistics

- Exact agreement: easiest to read; counts identical human labels.
- Cohen's kappa: corrects exact agreement for chance agreement.
- Quadratic weighted kappa: useful because `{TACNO}`, `{BLIZU}`, `{NETACNO}` and sentiment intensity values are ordered.
- Binary acceptable agreement: collapses `{TACNO}` + `{BLIZU}` vs `{NETACNO}` for the practical question "is the synthetic label usable?".
- Effective value agreement: compares the final human sentiment value after applying corrections from `human_*_value`.

## Row-level agreement

| Metric | Count | Rate |
|---|---:|---:|
| Both axes have same status | {row["status_both_axes_agree_count"]} / {stats["common_rows"]} | {row["status_both_axes_agree_pct"]}% |
| Both axes have same effective value | {row["effective_values_both_axes_agree_count"]} / {stats["common_rows"]} | {row["effective_values_both_axes_agree_pct"]}% |
| Both axes agree on acceptable/not acceptable | {row["acceptable_both_axes_agree_count"]} / {stats["common_rows"]} | {row["acceptable_both_axes_agree_pct"]}% |

## Status agreement

| Scope | Exact agreement | Cohen kappa | Weighted kappa |
|---|---:|---:|---:|
| All axes | {status["all_axes"]["exact_agreement_pct"]}% | {kappa_text(status["all_axes"]["cohen_kappa"])} | {kappa_text(status["all_axes"]["quadratic_weighted_kappa"])} |
| Positive | {status["positive"]["exact_agreement_pct"]}% | {kappa_text(status["positive"]["cohen_kappa"])} | {kappa_text(status["positive"]["quadratic_weighted_kappa"])} |
| Negative | {status["negative"]["exact_agreement_pct"]}% | {kappa_text(status["negative"]["cohen_kappa"])} | {kappa_text(status["negative"]["quadratic_weighted_kappa"])} |

## Acceptable binary agreement

| Scope | Exact agreement | Cohen kappa |
|---|---:|---:|
| All axes | {acceptable["all_axes"]["exact_agreement_pct"]}% | {kappa_text(acceptable["all_axes"]["cohen_kappa"])} |
| Positive | {acceptable["positive"]["exact_agreement_pct"]}% | {kappa_text(acceptable["positive"]["cohen_kappa"])} |
| Negative | {acceptable["negative"]["exact_agreement_pct"]}% | {kappa_text(acceptable["negative"]["cohen_kappa"])} |

## Effective value agreement

| Axis | Exact agreement | Within one step | Mean distance | Weighted kappa |
|---|---:|---:|---:|---:|
| Positive | {values["positive"]["exact_agreement_pct"]}% | {values["positive"]["within_one_step_pct"]}% | {values["positive"]["mean_absolute_distance"]} | {kappa_text(values["positive"]["quadratic_weighted_kappa"])} |
| Negative | {values["negative"]["exact_agreement_pct"]}% | {values["negative"]["within_one_step_pct"]}% | {values["negative"]["mean_absolute_distance"]} | {kappa_text(values["negative"]["quadratic_weighted_kappa"])} |

## Where they differ

Rows with complete agreement on status, binary acceptability, and effective values: {len(disagreements["full_agreement_rows"])} / {stats["common_rows"]}.
Rows with any disagreement: {len(disagreements["any_disagreement_rows"])} / {stats["common_rows"]}.

| Difference type | Count |
|---|---:|
| Positive status | {len(disagreements["by_type"]["positive_status"])} |
| Negative status | {len(disagreements["by_type"]["negative_status"])} |
| Positive acceptable/not acceptable | {len(disagreements["by_type"]["positive_acceptable"])} |
| Negative acceptable/not acceptable | {len(disagreements["by_type"]["negative_acceptable"])} |
| Positive effective value | {len(disagreements["by_type"]["positive_effective_value"])} |
| Negative effective value | {len(disagreements["by_type"]["negative_effective_value"])} |

Binary acceptability disagreements are the most important practical conflicts:

- Positive: {", ".join(disagreements["by_type"]["positive_acceptable"]) or "none"}
- Negative: {", ".join(disagreements["by_type"]["negative_acceptable"]) or "none"}

`comparison_{label_a}_{label_b}.csv` lists every aligned row. `disagreements_{label_a}_{label_b}.csv` lists rows where the annotators differ by status, acceptability, or effective value.
"""
    if stats["only_in_first"] or stats["only_in_second"]:
        text += "\n## Alignment warnings\n\n"
        if stats["only_in_first"]:
            text += f"- Only in {label_a}: {', '.join(stats['only_in_first'])}\n"
        if stats["only_in_second"]:
            text += f"- Only in {label_b}: {', '.join(stats['only_in_second'])}\n"
    return text


def fleiss_kappa(subject_ratings: list[list[str]], labels: list[str]) -> float | None:
    if not subject_ratings:
        return None
    raters = len(subject_ratings[0])
    if raters < 2:
        return None
    usable = [ratings for ratings in subject_ratings if len(ratings) == raters]
    if not usable:
        return None

    p_bar = 0.0
    label_totals = Counter()
    for ratings in usable:
        counts = Counter(ratings)
        label_totals.update(counts)
        p_bar += (sum(count * count for count in counts.values()) - raters) / (
            raters * (raters - 1)
        )
    p_bar /= len(usable)
    total_ratings = len(usable) * raters
    p_e = sum((label_totals[label] / total_ratings) ** 2 for label in labels)
    if p_e == 1:
        return None
    return round((p_bar - p_e) / (1 - p_e), 4)


def majority_value(values: list[str]) -> str:
    counts = Counter(values)
    value, count = counts.most_common(1)[0]
    return value if count > len(values) / 2 else ""


def multi_agreement_stats(subject_ratings: list[list[str]], labels: list[str]) -> dict[str, Any]:
    total = len(subject_ratings)
    complete = sum(1 for ratings in subject_ratings if len(set(ratings)) == 1)
    with_majority = sum(1 for ratings in subject_ratings if majority_value(ratings))
    return {
        "n": total,
        "complete_agreement_count": complete,
        "complete_agreement_pct": pct(complete, total),
        "majority_count": with_majority,
        "majority_pct": pct(with_majority, total),
        "fleiss_kappa": fleiss_kappa(subject_ratings, labels),
    }


def compare_three_stats(labels: list[str], out_root: Path) -> dict[str, Any]:
    if len(labels) < 3:
        raise ValueError("compare_three_stats requires at least three labels")

    row_maps: dict[str, dict[str, dict[str, Any]]] = {}
    ordered_first_ili: list[str] = []
    for index, label in enumerate(labels):
        workbook = out_root / label / f"human_eval_{label}.xlsx"
        _, rows = load_rows(workbook)
        row_maps[label] = {row["ILI"]: row for row in rows}
        if index == 0:
            ordered_first_ili = [row["ILI"] for row in rows]

    common = set(row_maps[labels[0]])
    for label in labels[1:]:
        common &= set(row_maps[label])
    common_ili = [ili for ili in ordered_first_ili if ili in common]
    only_by_label = {
        label: sorted(set(row_maps[label]) - common)
        for label in labels
        if sorted(set(row_maps[label]) - common)
    }

    comparison_name = "_".join(labels)
    out_dir = out_root / f"comparison_{comparison_name}"
    out_dir.mkdir(parents=True, exist_ok=True)
    all_csv = out_dir / f"comparison_{comparison_name}.csv"
    disagreement_csv = out_dir / f"disagreements_{comparison_name}.csv"
    out_json = out_dir / f"comparison_{comparison_name}_stats.json"
    out_md = out_dir / "README.md"

    status_subjects = {"all_axes": [], "positive": [], "negative": []}
    acceptable_subjects = {"all_axes": [], "positive": [], "negative": []}
    value_subjects = {"positive": [], "negative": []}
    value_pair_distances = {"positive": [], "negative": []}
    detail_rows: list[dict[str, Any]] = []

    row_status_all_agree = 0
    row_values_all_agree = 0
    row_acceptable_all_agree = 0

    for ili in common_ili:
        base_row = row_maps[labels[0]][ili]
        detail: dict[str, Any] = {
            "ILI": ili,
            "lemma_names": base_row.get("lemma_names"),
            "definition": base_row.get("definition"),
            "sentiment_lexicon": base_row.get("sentiment_lexicon"),
        }
        axis_status_agree: list[bool] = []
        axis_value_agree: list[bool] = []
        axis_acceptable_agree: list[bool] = []

        for axis in ["positive", "negative"]:
            statuses = [
                norm(row_maps[label][ili].get(f"human_{axis}_eval")) for label in labels
            ]
            acceptables = [
                "acceptable" if status in (TACNO, BLIZU) else "not_acceptable"
                for status in statuses
            ]
            values = [
                effective_value(row_maps[label][ili], axis) for label in labels
            ]
            scores = VALUE_SCORES[axis]

            status_subjects[axis].append(statuses)
            status_subjects["all_axes"].append(statuses)
            acceptable_subjects[axis].append(acceptables)
            acceptable_subjects["all_axes"].append(acceptables)
            value_subjects[axis].append(values)

            distances = []
            for left, right in combinations(values, 2):
                if left in scores and right in scores:
                    distance = abs(scores[left] - scores[right])
                    distances.append(distance)
                    value_pair_distances[axis].append(distance)

            status_agree = len(set(statuses)) == 1
            acceptable_agree = len(set(acceptables)) == 1
            value_agree = len(set(values)) == 1
            axis_status_agree.append(status_agree)
            axis_acceptable_agree.append(acceptable_agree)
            axis_value_agree.append(value_agree)

            for label, status, acceptable, value in zip(labels, statuses, acceptables, values):
                detail[f"{axis}_status_{label}"] = status
                detail[f"{axis}_acceptable_{label}"] = acceptable
                detail[f"{axis}_value_{label}"] = value
            detail[f"{axis}_status_all_agree"] = status_agree
            detail[f"{axis}_status_majority"] = majority_value(statuses)
            detail[f"{axis}_acceptable_all_agree"] = acceptable_agree
            detail[f"{axis}_acceptable_majority"] = majority_value(acceptables)
            detail[f"{axis}_value_all_agree"] = value_agree
            detail[f"{axis}_value_majority"] = majority_value(values)
            detail[f"{axis}_max_pair_distance"] = max(distances) if distances else None

        detail["row_status_both_axes_all_agree"] = all(axis_status_agree)
        detail["row_values_both_axes_all_agree"] = all(axis_value_agree)
        detail["row_acceptable_both_axes_all_agree"] = all(axis_acceptable_agree)
        row_status_all_agree += int(detail["row_status_both_axes_all_agree"])
        row_values_all_agree += int(detail["row_values_both_axes_all_agree"])
        row_acceptable_all_agree += int(detail["row_acceptable_both_axes_all_agree"])
        detail_rows.append(detail)

    fieldnames = list(detail_rows[0].keys()) if detail_rows else []
    disagreements = [
        row
        for row in detail_rows
        if not row["row_status_both_axes_all_agree"]
        or not row["row_values_both_axes_all_agree"]
        or not row["row_acceptable_both_axes_all_agree"]
    ]
    for csv_path, rows in ((all_csv, detail_rows), (disagreement_csv, disagreements)):
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    status_agreement = {
        key: multi_agreement_stats(subjects, STATUS_ORDER)
        for key, subjects in status_subjects.items()
    }
    acceptable_agreement = {
        key: multi_agreement_stats(subjects, ["not_acceptable", "acceptable"])
        for key, subjects in acceptable_subjects.items()
    }

    value_agreement: dict[str, Any] = {}
    for axis, value_labels in (("positive", POSITIVE_VALUES), ("negative", NEGATIVE_VALUES)):
        subjects = value_subjects[axis]
        complete = sum(1 for values in subjects if len(set(values)) == 1)
        within_one_step = 0
        for values in subjects:
            scores = VALUE_SCORES[axis]
            distances = [
                abs(scores[left] - scores[right])
                for left, right in combinations(values, 2)
                if left in scores and right in scores
            ]
            if distances and max(distances) <= 0.25:
                within_one_step += 1
        pairwise_kappas = {}
        for left_label, right_label in combinations(labels, 2):
            left_values = [
                effective_value(row_maps[left_label][ili], axis) for ili in common_ili
            ]
            right_values = [
                effective_value(row_maps[right_label][ili], axis) for ili in common_ili
            ]
            pairwise_kappas[f"{left_label}_vs_{right_label}"] = weighted_kappa(
                left_values, right_values, value_labels, quadratic=True
            )
        numeric_kappas = [
            value for value in pairwise_kappas.values() if value is not None
        ]
        distances = value_pair_distances[axis]
        value_agreement[axis] = {
            "n": len(subjects),
            "complete_agreement_count": complete,
            "complete_agreement_pct": pct(complete, len(subjects)),
            "all_pairwise_within_one_step_count": within_one_step,
            "all_pairwise_within_one_step_pct": pct(within_one_step, len(subjects)),
            "mean_pairwise_distance": round(sum(distances) / len(distances), 4)
            if distances
            else None,
            "max_pairwise_distance": max(distances) if distances else None,
            "pairwise_quadratic_weighted_kappa": pairwise_kappas,
            "mean_pairwise_quadratic_weighted_kappa": round(
                sum(numeric_kappas) / len(numeric_kappas), 4
            )
            if numeric_kappas
            else None,
        }

    disagreement_lists = {
        "positive_status": [
            row["ILI"] for row in detail_rows if not row["positive_status_all_agree"]
        ],
        "negative_status": [
            row["ILI"] for row in detail_rows if not row["negative_status_all_agree"]
        ],
        "positive_acceptable": [
            row["ILI"] for row in detail_rows if not row["positive_acceptable_all_agree"]
        ],
        "negative_acceptable": [
            row["ILI"] for row in detail_rows if not row["negative_acceptable_all_agree"]
        ],
        "positive_effective_value": [
            row["ILI"] for row in detail_rows if not row["positive_value_all_agree"]
        ],
        "negative_effective_value": [
            row["ILI"] for row in detail_rows if not row["negative_value_all_agree"]
        ],
    }
    any_disagreement = sorted(
        set().union(*(set(values) for values in disagreement_lists.values()))
    )
    full_agreement = [
        row["ILI"]
        for row in detail_rows
        if row["row_status_both_axes_all_agree"]
        and row["row_values_both_axes_all_agree"]
        and row["row_acceptable_both_axes_all_agree"]
    ]

    stats = {
        "comparison": comparison_name,
        "labels": labels,
        "common_rows": len(common_ili),
        "only_by_label": only_by_label,
        "row_level": {
            "status_both_axes_all_agree_count": row_status_all_agree,
            "status_both_axes_all_agree_pct": pct(row_status_all_agree, len(common_ili)),
            "effective_values_both_axes_all_agree_count": row_values_all_agree,
            "effective_values_both_axes_all_agree_pct": pct(
                row_values_all_agree, len(common_ili)
            ),
            "acceptable_both_axes_all_agree_count": row_acceptable_all_agree,
            "acceptable_both_axes_all_agree_pct": pct(
                row_acceptable_all_agree, len(common_ili)
            ),
        },
        "status_agreement": status_agreement,
        "acceptable_binary_agreement": acceptable_agreement,
        "effective_value_agreement": value_agreement,
        "disagreement_summary": {
            "full_agreement_rows": full_agreement,
            "any_disagreement_rows": any_disagreement,
            "by_type": disagreement_lists,
        },
        "comparison_csv": str(all_csv),
        "disagreement_csv": str(disagreement_csv),
    }
    out_json.write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_md.write_text(compare_three_readme(labels, stats), encoding="utf-8")
    return stats


def compare_three_readme(labels: list[str], stats: dict[str, Any]) -> str:
    row = stats["row_level"]
    status = stats["status_agreement"]
    acceptable = stats["acceptable_binary_agreement"]
    values = stats["effective_value_agreement"]
    disagreements = stats["disagreement_summary"]
    comparison_name = "_".join(labels)
    text = f"""# Human evaluation comparison {comparison_name}

Compared rows: {stats["common_rows"]}

## Recommended agreement statistics

- Complete agreement: all annotators chose the same label.
- Fleiss' kappa: chance-corrected agreement for all annotators.
- Binary acceptable agreement: collapses `{TACNO}` + `{BLIZU}` vs `{NETACNO}`.
- Effective value agreement: compares the final human sentiment value after corrections.

## Row-level complete agreement

| Metric | Count | Rate |
|---|---:|---:|
| Both axes have same status across all annotators | {row["status_both_axes_all_agree_count"]} / {stats["common_rows"]} | {row["status_both_axes_all_agree_pct"]}% |
| Both axes have same effective value across all annotators | {row["effective_values_both_axes_all_agree_count"]} / {stats["common_rows"]} | {row["effective_values_both_axes_all_agree_pct"]}% |
| Both axes agree on acceptable/not acceptable across all annotators | {row["acceptable_both_axes_all_agree_count"]} / {stats["common_rows"]} | {row["acceptable_both_axes_all_agree_pct"]}% |

## Status agreement

| Scope | Complete agreement | Majority exists | Fleiss kappa |
|---|---:|---:|---:|
| All axes | {status["all_axes"]["complete_agreement_pct"]}% | {status["all_axes"]["majority_pct"]}% | {kappa_text(status["all_axes"]["fleiss_kappa"])} |
| Positive | {status["positive"]["complete_agreement_pct"]}% | {status["positive"]["majority_pct"]}% | {kappa_text(status["positive"]["fleiss_kappa"])} |
| Negative | {status["negative"]["complete_agreement_pct"]}% | {status["negative"]["majority_pct"]}% | {kappa_text(status["negative"]["fleiss_kappa"])} |

## Acceptable binary agreement

| Scope | Complete agreement | Majority exists | Fleiss kappa |
|---|---:|---:|---:|
| All axes | {acceptable["all_axes"]["complete_agreement_pct"]}% | {acceptable["all_axes"]["majority_pct"]}% | {kappa_text(acceptable["all_axes"]["fleiss_kappa"])} |
| Positive | {acceptable["positive"]["complete_agreement_pct"]}% | {acceptable["positive"]["majority_pct"]}% | {kappa_text(acceptable["positive"]["fleiss_kappa"])} |
| Negative | {acceptable["negative"]["complete_agreement_pct"]}% | {acceptable["negative"]["majority_pct"]}% | {kappa_text(acceptable["negative"]["fleiss_kappa"])} |

## Effective value agreement

| Axis | Complete agreement | All pairs within one step | Mean pairwise distance | Mean pairwise weighted kappa |
|---|---:|---:|---:|---:|
| Positive | {values["positive"]["complete_agreement_pct"]}% | {values["positive"]["all_pairwise_within_one_step_pct"]}% | {values["positive"]["mean_pairwise_distance"]} | {values["positive"]["mean_pairwise_quadratic_weighted_kappa"]} |
| Negative | {values["negative"]["complete_agreement_pct"]}% | {values["negative"]["all_pairwise_within_one_step_pct"]}% | {values["negative"]["mean_pairwise_distance"]} | {values["negative"]["mean_pairwise_quadratic_weighted_kappa"]} |

## Where they differ

Rows with complete agreement on status, binary acceptability, and effective values across all annotators: {len(disagreements["full_agreement_rows"])} / {stats["common_rows"]}.
Rows with any disagreement: {len(disagreements["any_disagreement_rows"])} / {stats["common_rows"]}.

| Difference type | Count |
|---|---:|
| Positive status | {len(disagreements["by_type"]["positive_status"])} |
| Negative status | {len(disagreements["by_type"]["negative_status"])} |
| Positive acceptable/not acceptable | {len(disagreements["by_type"]["positive_acceptable"])} |
| Negative acceptable/not acceptable | {len(disagreements["by_type"]["negative_acceptable"])} |
| Positive effective value | {len(disagreements["by_type"]["positive_effective_value"])} |
| Negative effective value | {len(disagreements["by_type"]["negative_effective_value"])} |

Binary acceptability disagreements:

- Positive: {", ".join(disagreements["by_type"]["positive_acceptable"]) or "none"}
- Negative: {", ".join(disagreements["by_type"]["negative_acceptable"]) or "none"}

`comparison_{comparison_name}.csv` lists every aligned row. `disagreements_{comparison_name}.csv` lists rows where at least one agreement check differs.
"""
    if stats["only_by_label"]:
        text += "\n## Alignment warnings\n\n"
        for label, values in stats["only_by_label"].items():
            text += f"- Only in {label}: {', '.join(values)}\n"
    return text


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--eval",
        action="append",
        nargs=2,
        metavar=("LABEL", "XLSX"),
        help="Evaluation label and source workbook path.",
    )
    parser.add_argument("--compare", nargs=2, metavar=("LABEL_A", "LABEL_B"))
    parser.add_argument("--compare-three", nargs="+", metavar="LABEL")
    parser.add_argument("--out-root", default="human_eval")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    out_root = Path(args.out_root)
    evals = args.eval
    compare = args.compare
    compare_three = args.compare_three
    if not evals:
        candidates = [
            ("01", "C:/Users/sasa5/Downloads/PZL_tre\u0107i zadatak.xlsx"),
            ("02", "C:/Users/sasa5/Downloads/PZL_tre\u0107i zadatak(1).xlsx"),
            ("03", "C:/Users/sasa5/Downloads/PZL_tre\u0107i zadatak_Sofija.xlsx"),
        ]
        evals = []
        available_labels = []
        for label, workbook in candidates:
            copied_workbook = out_root / label / f"human_eval_{label}.xlsx"
            if Path(workbook).exists():
                evals.append((label, workbook))
                available_labels.append(label)
            elif copied_workbook.exists():
                available_labels.append(label)
        if len(evals) >= 3 and compare_three is None:
            compare_three = available_labels[:3]
        elif len(available_labels) >= 3 and compare_three is None:
            compare_three = available_labels[:3]
        elif compare is None and {"01", "02"} <= set(available_labels):
            compare = ("01", "02")

    summaries = []
    for label, workbook in evals:
        source = Path(workbook)
        if not source.exists():
            raise FileNotFoundError(source)
        summaries.append(single_stats(label, source, out_root))

    comparisons = {}
    if compare_three and compare is None:
        for label_a, label_b in combinations(compare_three, 2):
            comparison_key = f"{label_a}_{label_b}"
            comparisons[comparison_key] = compare_stats(label_a, label_b, out_root)
    elif compare:
        comparison_key = f"{compare[0]}_{compare[1]}"
        comparisons[comparison_key] = compare_stats(compare[0], compare[1], out_root)
    three_way = compare_three_stats(compare_three, out_root) if compare_three else None
    print(
        json.dumps(
            {
                "evaluations": [
                    {
                        "label": summary["label"],
                        "rows": summary["rows"],
                        "axis_total_eval_counts": summary["axis_total_eval_counts"],
                        "row_quality_counts": summary["row_quality_counts"],
                        "quality_flags": summary["quality_flags"],
                    }
                    for summary in summaries
                ],
                "comparisons": {
                    key: value["row_level"] for key, value in comparisons.items()
                },
                "three_way": three_way["row_level"] if three_way else None,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
